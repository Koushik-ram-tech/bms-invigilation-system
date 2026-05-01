"""
BMS Exam Invigilation System v10 — Final Backend
=================================================
Changes v10 (over v9):
  ✦ ROOMS ENDPOINT: GET/POST /rooms — view and update room capacities at any time.
    Changes persist for the session. Re-running allocation uses updated room config.
  ✦ NAME AUTO-FILL FIX: student_list now stored with normalised USN keys (upper,
    stripped). Attendance sheet lookup also normalises the USN key before lookup.
    Names now correctly appear regardless of year or USN format.
  ✦ UNIVERSAL USN: works for any year/batch — 1BM25MC, 1BM26MC, 1BM23MCA, etc.
    Room USN ranges rebuilt automatically from student list's detected prefix.
  ✦ All v9 features retained: backup invigilators, duplicate fix, reset fix.

Run:  uvicorn main:app --reload --host 0.0.0.0 --port 8000
Deps: pip install fastapi uvicorn[standard] python-multipart
      pdfplumber openpyxl python-docx pandas
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Response, Query
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Optional
from datetime import datetime
from collections import defaultdict
import io, re, statistics
import os
import threading
from copy import deepcopy
from utils import normalise_usn, rebuild_rooms_with_prefix, log_error, log_info
from fastapi import Request
from fastapi.responses import JSONResponse
import json

from bms_parser import (
    parse_timetable, parse_exam_schedule, parse_student_list,
    parse_date, parse_time, fmt_date, fmt_time,
    DEFAULT_ROOMS, DEFAULT_STUDENTS, KNOWN_FACULTY,
    generate_usns,
)


ROOM_FILE = "config/rooms.json"

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── App ────────────────────────────────────────────────────────
app = FastAPI(
    title="BMS Exam Invigilation System v10",
    description="Upload timetable + exam schedule → auto-allocated duty roster with backups",
    version="10.0.0",
)



@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    log_error(f"Unhandled error: {str(exc)}")

    return JSONResponse(
        status_code=500,
        content={
            "status": "error",
            "message": "Internal server error",
            "detail": str(exc)[:200]
        }
    )

ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "ALLOWED_ORIGINS",
        "http://localhost:3000,http://127.0.0.1:3000,http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS, allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ── Global state ───────────────────────────────────────────────
faculty_map:    Dict[str, str]  = {}
busy_slots:     List[Dict]      = []
exam_schedule:  List[Dict]      = []
duty_records:   List[Dict]      = []
# KEY FIX: student_list keys are ALWAYS normalise_usn(usn) — upper-stripped
student_list:   Dict[str, Dict] = {}
# Room config — editable via /rooms endpoint
room_config:    List[Dict]      = deepcopy(DEFAULT_ROOMS)
# Detected USN prefix from uploaded student list
detected_usn_prefix: str        = ""

timetable_done = False
exam_done      = False
students_done  = False
multi_batch    = False
last_result:   Dict = {}
state_lock = threading.RLock()

ACCEPTED_EXTS = {"pdf", "docx", "doc", "xlsx", "xls", "ods", "csv"}
NUM_BACKUPS   = 2


# ══════════════════════════════════════════════════════════════
#  ROOMS HELPERS
# ══════════════════════════════════════════════════════════════

def _current_rooms() -> List[Dict]:
    """Return the currently active room configuration (deep copy)."""
    return deepcopy(room_config)


def _rooms_with_usn_ranges() -> List[Dict]:
    """
    Return room_config with USN ranges computed from detected_usn_prefix.
    If no prefix detected yet, returns room_config as-is.
    """
    if not detected_usn_prefix:
        return _current_rooms()
    total   = sum(r["students"] for r in room_config)
    rebuilt = rebuild_rooms_with_prefix(_current_rooms(), detected_usn_prefix, total)
    return rebuilt


# ══════════════════════════════════════════════════════════════
#  ALLOCATION HELPERS
# ══════════════════════════════════════════════════════════════

def _duty_count(name: str) -> int:
    return sum(1 for r in duty_records if r.get("faculty_name") == name)


def _days_since_last(name: str, exam_date: str) -> float:
    recs = [r for r in duty_records if r.get("faculty_name") == name]
    if not recs: return 999.0
    exam_dt = parse_date(exam_date)
    if not exam_dt: return 0.0
    dates = [parse_date(r["exam_date"]) for r in recs]
    valid = [d for d in dates if d]
    return max(0.0, (exam_dt - max(valid)).days) if valid else 0.0


def _teaches(code: str, subject: str) -> bool:
    exam_words = {w.lower() for w in subject.split() if len(w) > 3}
    for slot in busy_slots:
        if slot.get("faculty_code", "").upper() != code.upper(): continue
        slot_words = {w.lower() for w in slot.get("subject", "").split() if len(w) > 3}
        if exam_words & slot_words: return True
    return False


def _is_teaching(code: str, day: str, start_time: str) -> bool:
    """Only fires when multi_batch=True (junior batches present)."""
    if not multi_batch: return False
    exam_start = parse_time(start_time)
    for slot in busy_slots:
        if slot.get("faculty_code", "").upper() != code.upper(): continue
        slot_day = slot.get("day", "").strip().lower()
        exam_day = day.strip().lower()
        if not slot_day or not exam_day: continue
        if slot_day[:3] != exam_day[:3]: continue
        st_raw = slot.get("time_slot", "")
        if not st_raw or not exam_start: return True
        m = re.search(r"[\d.]+", st_raw)
        if m:
            slot_dt = parse_time(m.group(0))
            if slot_dt and abs((exam_start - slot_dt).total_seconds()) / 3600 < 1.5:
                return True
    return False


def _already_at(exam_date: str, start_time: str) -> set:
    return {r["faculty_name"] for r in duty_records
            if r.get("exam_date") == exam_date and r.get("start_time") == start_time}


def _consecutive(name: str, exam_date: str, start_time: str) -> bool:
    es = parse_time(start_time)
    if not es: return False
    for r in duty_records:
        if r.get("faculty_name") != name or r.get("exam_date") != exam_date: continue
        pe = parse_time(r.get("end_time", ""))
        if pe and 0 < (es - pe).total_seconds() / 60 < 30: return True
    return False


def _score(name: str, exam_date: str) -> float:
    count = _duty_count(name); gap = _days_since_last(name, exam_date)
    return round(0.6 / (1.0 + count) + 0.4 * min(1.0, gap / 14.0), 4)


def _build_reasoning(name, code, subject, exam_date, day, start_time, rank, n) -> Dict:
    count = _duty_count(name); gap = _days_since_last(name, exam_date)
    score = _score(name, exam_date)
    vals  = [_duty_count(nm) for nm in faculty_map.values() if nm]
    avg   = round(statistics.mean(vals), 1) if vals else 0
    why   = []
    if count == 0:        why.append("Zero duties so far — top priority")
    elif count < avg:     why.append(f"Below-average load: {count} duties vs avg {avg}")
    else:                 why.append(f"All lighter-load faculty unavailable ({count} duties, avg {avg})")
    if gap >= 14:         why.append("14+ days since last duty")
    elif gap >= 7:        why.append(f"{int(gap)} days since last duty — adequate rest")
    checks = [
        f"Free on {day} — no class at {start_time}",
        f"Does not teach '{subject}' — no COI",
        "No consecutive duty — fatigue rule OK",
        f"Ranked #{rank} of {n} available",
    ]
    return {"selected_faculty":name,"faculty_code":code,"score":score,"rank":rank,
            "total_candidates":n,"why_selected":why,"checks_passed":checks,
            "duty_count":count,"avg_duty_count":avg,
            "summary":f"{name} — {count} duties (avg {avg}), {int(gap)} days rest, score {score}"}


def _pick_backups(primary_names: set, subject: str, exam_date: str,
                  day: str, start_time: str, count: int = NUM_BACKUPS) -> List[Dict]:
    active_map = faculty_map if faculty_map else KNOWN_FACULTY
    at_slot    = _already_at(exam_date, start_time)
    backups    = []
    for code, name in active_map.items():
        if not name or name in primary_names or name in at_slot: continue
        if _teaches(code, subject) or _is_teaching(code, day, start_time): continue
        if _consecutive(name, exam_date, start_time): continue
        backups.append({"name":name,"code":code,"duties":_duty_count(name),"score":_score(name,exam_date)})
    backups.sort(key=lambda x: (x["duties"], -x["score"]))
    return [{"faculty_name":b["name"],"faculty_code":b["code"],"current_duties":b["duties"],
             "score":b["score"],"note":"Backup — not counted in workload unless activated"}
            for b in backups[:count]]


def _intelligence() -> Dict:
    if not duty_records: return {"available": False}
    counts  = {n: _duty_count(n) for n in faculty_map.values() if n}
    never   = [n for n, c in counts.items() if c == 0]
    day_cnt = defaultdict(int)
    for r in duty_records: day_cnt[r.get("day","?")] += 1
    vals    = list(counts.values())
    mean    = round(statistics.mean(vals), 1) if vals else 0
    std     = round(statistics.stdev(vals), 2) if len(vals) > 1 else 0
    top3    = sorted(counts.items(), key=lambda x: -x[1])[:3]
    busiest = max(day_cnt, key=day_cnt.get) if day_cnt else "N/A"
    return {"available":True,"total_duties":len(duty_records),
            "total_exams":len({(r.get("subject"),r.get("exam_date")) for r in duty_records}),
            "zero_duty_faculty":never,"zero_count":len(never),"busiest_day":busiest,
            "mean_duties":mean,"std_duties":std,
            "top_invigilators":[{"name":n,"duties":c} for n,c in top3],
            "suggestion":(f"{len(never)} faculty member(s) have zero duties." if never
                          else "Good distribution." if std < 1.5 else "Uneven distribution — consider re-running.")}


def _allocate_one(exam: Dict) -> Dict:
    subject   = exam.get("subject","Unknown")
    exam_date = exam.get("exam_date","")
    day       = exam.get("day","")
    start_t   = exam.get("start_time","9:15 AM")
    end_t     = exam.get("end_time","10:30 AM")
    rooms     = exam.get("rooms", _rooms_with_usn_ranges())
    total_stu = exam.get("total_students", DEFAULT_STUDENTS)

    active_map = faculty_map if faculty_map else KNOWN_FACULTY
    double     = _already_at(exam_date, start_t)
    candidates, excluded = [], []

    for code, name in active_map.items():
        if not name: continue
        if name in double:
            excluded.append({"name":name,"reason":"Already assigned at this time"})
        elif _consecutive(name, exam_date, start_t):
            excluded.append({"name":name,"reason":"Consecutive duty — fatigue rule"})
        elif _teaches(code, subject):
            excluded.append({"name":name,"reason":f"Teaches {subject} (COI)"})
        elif _is_teaching(code, day, start_t):
            excluded.append({"name":name,"reason":f"Has class on {day} at {start_t}"})
        else:
            candidates.append({"name":name,"code":code,"score":_score(name,exam_date),"duties":_duty_count(name)})

    candidates.sort(key=lambda x: -x["score"])
    halls = len(rooms)
    if len(candidates) < halls:
        if not candidates:
            return {"success":False,"subject":subject,"exam_date":exam_date,"start_time":start_t,"end_time":end_t,
                    "error":f"No invigilators available — all {len(excluded)} excluded.","excluded":excluded,"backups":[]}
        halls = len(candidates); rooms = rooms[:halls]

    assignments   = []; primary_names = set()
    for i, room in enumerate(rooms):
        fac       = candidates[i]
        reasoning = _build_reasoning(fac["name"],fac["code"],subject,exam_date,day,start_t,i+1,len(candidates))
        record    = {"subject":subject,"exam_type":exam.get("exam_type","CIE"),
                     "course_code":exam.get("course_code",""),"section":exam.get("section",""),
                     "exam_date":exam_date,"day":day,"start_time":start_t,"end_time":end_t,
                     "faculty_name":fac["name"],"faculty_code":fac["code"],
                     "hall":room.get("room",f"Hall {i+1}"),"students":room.get("students",30),
                     "usn_range":room.get("usn_range",""),"usn_start":room.get("usn_start",""),
                     "usn_end":room.get("usn_end",""),"reasoning":reasoning}
        if fac["name"] not in _already_at(exam_date, start_t):
            duty_records.append(record); primary_names.add(fac["name"])
        assignments.append(record)

    backups = _pick_backups(primary_names, subject, exam_date, day, start_t, NUM_BACKUPS)
    return {"success":True,"subject":subject,"exam_type":exam.get("exam_type","CIE"),
            "course_code":exam.get("course_code",""),"exam_date":exam_date,"day":day,
            "start_time":start_t,"end_time":end_t,"total_students":total_stu,"halls":halls,
            "assignments":assignments,"backups":backups,"excluded":excluded,
            "used_defaults":exam.get("used_defaults",[])}


def _run_all() -> Dict:
    def sk(e):
        d = parse_date(e.get("exam_date","")); t = parse_time(e.get("start_time",""))
        return (d or datetime.min, t or datetime.min)
    results  = [_allocate_one(e) for e in sorted(exam_schedule, key=sk)]
    ok       = sum(1 for r in results if r.get("success"))
    workload = {n: _duty_count(n) for n in faculty_map.values() if n}
    vals     = list(workload.values())
    return {"total":len(results),"success_count":ok,"fail_count":len(results)-ok,"results":results,
            "workload":sorted([{"name":n,"duties":c} for n,c in workload.items()],key=lambda x:-x["duties"]),
            "mean_duties":round(statistics.mean(vals),1) if vals else 0,
            "message":f"{ok}/{len(results)} exams allocated."}


# ══════════════════════════════════════════════════════════════
#  ATTENDANCE SHEET GENERATOR
#  KEY FIX: student_list lookup uses normalise_usn(usn) as key
#           so names ALWAYS match regardless of case/spaces/year
# ══════════════════════════════════════════════════════════════

def generate_attendance_sheets(exam_records: List[Dict]) -> bytes:
    if not EXCEL_OK: raise HTTPException(500, "openpyxl not installed")
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    COLLEGE_FONT = Font(name="Times New Roman", bold=True, size=14)
    DEPT_FONT    = Font(name="Times New Roman", bold=True, size=12)
    INFO_FONT    = Font(name="Times New Roman", size=11)
    HEADER_FONT  = Font(name="Times New Roman", bold=True, size=11)
    DATA_FONT    = Font(name="Times New Roman", size=10)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left", vertical="center")
    thin         = Side(style="thin")
    BDR          = Border(left=thin, right=thin, top=thin, bottom=thin)
    HEADER_FILL  = PatternFill("solid", fgColor="1565C0")
    HEADER_TXT   = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")
    ALT_FILL     = PatternFill("solid", fgColor="EBF5FB")
    SUMMARY_FILL = PatternFill("solid", fgColor="FFF9C4")
    SUMMARY_HDR  = PatternFill("solid", fgColor="F57F17")
    SUMMARY_HFNT = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")

    sheet_count = 0
    for exam_rec in exam_records:
        if not exam_rec.get("success"): continue
        for assign in exam_rec.get("assignments", []):
            hall        = assign.get("hall","Hall")
            subject     = assign.get("subject","")
            exam_date   = assign.get("exam_date","")
            day         = assign.get("day","")
            start_time  = assign.get("start_time","")
            end_time    = assign.get("end_time","")
            exam_type   = assign.get("exam_type","CIE")
            course_code = assign.get("course_code","")
            invigilator = assign.get("faculty_name","")
            usn_start   = assign.get("usn_start","")
            usn_end     = assign.get("usn_end","")
            students    = assign.get("students", 30)
            backups     = exam_rec.get("backups", [])

            # Generate USN list for this hall
            if usn_start and usn_end:
                usn_list = generate_usns(usn_start, usn_end)
            elif detected_usn_prefix:
                # Use detected prefix from uploaded student list
                room_num = re.search(r'\d+', hall)
                rn       = int(room_num.group()) if room_num else 1
                # Calculate offset from room_config
                offset = 0
                for rc in room_config:
                    if rc["room"] == hall: break
                    offset += rc["students"]
                usn_list = [f"{detected_usn_prefix}{str(offset+i+1).zfill(3)}" for i in range(students)]
            else:
                # Absolute fallback — generate sequential with default prefix
                room_num = re.search(r'\d+', hall)
                offset   = (int(room_num.group()) - 1) * 38 if room_num else 0
                usn_list = [f"1BM25MC{str(i+1+offset).zfill(3)}" for i in range(students)]

            short_subj = re.sub(r'[^\w\s]','',subject)[:15].strip()
            sheet_name = f"{hall}_{short_subj}"[:31]
            base, counter = sheet_name, 1
            while sheet_name in [s.title for s in wb.worksheets]:
                sheet_name = f"{base[:28]}_{counter}"; counter += 1

            ws = wb.create_sheet(title=sheet_name)
            sheet_count += 1

            def mc(r1, c1, r2, c2):
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

            # ── Header rows ────────────────────────────────────
            mc(1,1,1,5); ws["A1"] = "B.M.S. COLLEGE OF ENGINEERING, BENGALURU – 560 019"
            ws["A1"].font = COLLEGE_FONT; ws["A1"].alignment = CENTER; ws.row_dimensions[1].height = 22
            mc(2,1,2,5); ws["A2"] = "(Autonomous Institute, Affiliated to VTU)"
            ws["A2"].font = Font(name="Times New Roman", italic=True, size=11); ws["A2"].alignment = CENTER
            mc(3,1,3,5); ws["A3"] = "Department of Computer Applications"
            ws["A3"].font = DEPT_FONT; ws["A3"].alignment = CENTER
            mc(4,1,4,5); ws["A4"] = f"ATTENDANCE SHEET — {exam_type.upper()} EXAMINATION"
            ws["A4"].font = DEPT_FONT; ws["A4"].alignment = CENTER
            ws.row_dimensions[5].height = 6

            # ── Exam info ──────────────────────────────────────
            for ridx, (label, value) in enumerate([
                ("Subject",     f"{course_code}  {subject}".strip()),
                ("Date & Day",  f"{exam_date}  ({day})"),
                ("Time",        f"{start_time}  to  {end_time}"),
                ("Hall / Room", hall),
                ("Invigilator", invigilator),
            ], 6):
                ws.cell(row=ridx, column=1, value=label).font = HEADER_FONT
                ws.cell(row=ridx, column=1).alignment = LEFT
                ws.cell(row=ridx, column=1).border = BDR
                mc(ridx,2,ridx,5)
                ws.cell(row=ridx, column=2, value=value).font = INFO_FONT
                ws.cell(row=ridx, column=2).alignment = LEFT
                ws.cell(row=ridx, column=2).border = BDR
                ws.row_dimensions[ridx].height = 18
            ws.row_dimensions[11].height = 6

            # ── Table header ───────────────────────────────────
            for col, hdr in enumerate(["Sl. No.","USN","Student Name","Programme","Signature"], 1):
                cell = ws.cell(row=12, column=col, value=hdr)
                cell.font = HEADER_TXT; cell.fill = HEADER_FILL
                cell.alignment = CENTER; cell.border = BDR
            ws.row_dimensions[12].height = 20

            # ── Student rows ───────────────────────────────────
            for idx, usn in enumerate(usn_list, 1):
                rn  = 12 + idx; alt = idx % 2 == 0
                # KEY FIX: always normalise USN before lookup
                key  = normalise_usn(usn)
                info = student_list.get(key, {})
                name = info.get("name", "")
                prog = info.get("programme", "MCA")

                ws.cell(row=rn, column=1, value=idx).alignment = CENTER
                ws.cell(row=rn, column=1).border = BDR; ws.cell(row=rn, column=1).font = DATA_FONT
                ws.cell(row=rn, column=2, value=usn).alignment = CENTER
                ws.cell(row=rn, column=2).border = BDR; ws.cell(row=rn, column=2).font = DATA_FONT
                ws.cell(row=rn, column=3, value=name).border = BDR
                ws.cell(row=rn, column=3).font = DATA_FONT; ws.cell(row=rn, column=3).alignment = LEFT
                ws.cell(row=rn, column=4, value=prog).alignment = CENTER
                ws.cell(row=rn, column=4).border = BDR; ws.cell(row=rn, column=4).font = DATA_FONT
                ws.cell(row=rn, column=5, value="").border = BDR
                if alt:
                    for col in range(1,6): ws.cell(row=rn, column=col).fill = ALT_FILL
                ws.row_dimensions[rn].height = 18

            # ── Footer: signatures ─────────────────────────────
            footer_row = 13 + len(usn_list) + 1
            ws.row_dimensions[footer_row].height = 6
            sig_row = footer_row + 1
            mc(sig_row,1,sig_row,2)
            ws.cell(row=sig_row,column=1,value="Invigilator Signature:").font = HEADER_FONT
            ws.cell(row=sig_row,column=1).border = BDR; ws.cell(row=sig_row,column=1).alignment = LEFT
            mc(sig_row,3,sig_row,5); ws.cell(row=sig_row,column=3,value="").border = BDR
            ws.row_dimensions[sig_row].height = 30
            hod_row = sig_row + 1
            mc(hod_row,1,hod_row,2)
            ws.cell(row=hod_row,column=1,value="HOD / Exam Coordinator:").font = HEADER_FONT
            ws.cell(row=hod_row,column=1).border = BDR; ws.cell(row=hod_row,column=1).alignment = LEFT
            mc(hod_row,3,hod_row,5); ws.cell(row=hod_row,column=3,value="").border = BDR
            ws.row_dimensions[hod_row].height = 30

            # ── Backup invigilators ────────────────────────────
            if backups:
                backup_spacer = hod_row + 1; ws.row_dimensions[backup_spacer].height = 8
                bhr = backup_spacer + 1; mc(bhr,1,bhr,5)
                bh = ws.cell(row=bhr,column=1,value="BACKUP INVIGILATORS (Standby — activate if primary absent)")
                bh.font = Font(name="Times New Roman",bold=True,size=11,color="FFFFFF")
                bh.fill = PatternFill("solid",fgColor="0D47A1"); bh.alignment = CENTER; bh.border = BDR
                ws.row_dimensions[bhr].height = 20
                for bi, bk in enumerate(backups, bhr+1):
                    mc(bi,1,bi,2)
                    ws.cell(row=bi,column=1,value=f"Backup {bi-bhr}: {bk['faculty_name']}").font = HEADER_FONT
                    ws.cell(row=bi,column=1).border = BDR; ws.cell(row=bi,column=1).alignment = LEFT
                    mc(bi,3,bi,3)
                    ws.cell(row=bi,column=3,value=f"Code: {bk['faculty_code']}  |  Duties: {bk['current_duties']}").font = DATA_FONT
                    ws.cell(row=bi,column=3).border = BDR
                    mc(bi,4,bi,5); ws.cell(row=bi,column=4,value="Signature:").font = DATA_FONT
                    ws.cell(row=bi,column=4).border = BDR; ws.row_dimensions[bi].height = 22
                next_row = bhr + 1 + len(backups) + 1
            else:
                next_row = hod_row + 2

            # ── Summary table (3 cols) ─────────────────────────
            ws.row_dimensions[next_row-1].height = 10
            shr = next_row; col_spans = [(1,1),(2,3),(4,5)]
            for (c1,c2), heading in zip(col_spans, ["No. of Absentees","No. of Booklets Collected",""]):
                if c1 != c2: mc(shr,c1,shr,c2)
                cell = ws.cell(row=shr,column=c1,value=heading)
                cell.font=SUMMARY_HFNT; cell.fill=SUMMARY_HDR; cell.alignment=CENTER; cell.border=BDR
            ws.row_dimensions[shr].height = 22
            sdr = shr + 1
            for (c1,c2) in col_spans:
                if c1 != c2: mc(sdr,c1,sdr,c2)
                cell = ws.cell(row=sdr,column=c1,value="")
                cell.fill=SUMMARY_FILL; cell.alignment=CENTER; cell.border=BDR
                cell.font=Font(name="Times New Roman",size=12)
            ws.row_dimensions[sdr].height = 36

            ws.column_dimensions["A"].width = 8;  ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 32; ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 22
            ws.page_setup.orientation="portrait"; ws.page_setup.paperSize=9
            ws.page_setup.fitToPage=True; ws.page_setup.fitToHeight=0; ws.page_setup.fitToWidth=1
            ws.print_title_rows="1:12"
            ws.page_margins.left=0.5; ws.page_margins.right=0.5
            ws.page_margins.top=0.75; ws.page_margins.bottom=0.75

    if sheet_count == 0:
        ws = wb.create_sheet("No Data"); ws["A1"] = "No attendance sheets generated. Run allocation first."
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _build_duty_excel() -> bytes:
    if not EXCEL_OK: raise HTTPException(500, "openpyxl not installed")
    wb  = openpyxl.Workbook()
    BLU = PatternFill("solid", fgColor="1565C0"); BFT = Font(color="FFFFFF",bold=True,size=11)
    ALT = PatternFill("solid", fgColor="E3F2FD"); th = Side(style="thin",color="CCCCCC")
    BDR = Border(left=th,right=th,top=th,bottom=th)
    CTR = Alignment(horizontal="center",vertical="center",wrap_text=True)
    def hr(ws,n):
        for c in ws[n]: c.fill=BLU; c.font=BFT; c.alignment=CTR; c.border=BDR
    def dr(ws,n,alt=False):
        for c in ws[n]: c.border=BDR; c.alignment=CTR
        if alt:
            for c in ws[n]: c.fill=ALT
    def sk(r):
        d=parse_date(r.get("exam_date","")); t=parse_time(r.get("start_time",""))
        return (d or datetime.min, t or datetime.min)
    ws1=wb.active; ws1.title="Invigilator Duty Chart"
    ws1.merge_cells("A1:L1"); ws1["A1"]="BMS College of Engineering — Invigilator Duty Allocation Chart"
    ws1["A1"].font=Font(bold=True,size=13); ws1["A1"].alignment=CTR; ws1.row_dimensions[1].height=24
    ws1.append([]); ws1.append(["Type","Date","Day","Subject","Code","Section","Start","End","Hall","Invigilator","Students","USN Range"])
    hr(ws1,3)
    for idx,r in enumerate(sorted(duty_records,key=sk),4):
        ws1.append([r.get("exam_type",""),r.get("exam_date",""),r.get("day",""),r.get("subject",""),
                    r.get("course_code",""),r.get("section",""),r.get("start_time",""),r.get("end_time",""),
                    r.get("hall",""),r.get("faculty_name",""),r.get("students",""),r.get("usn_range","")])
        dr(ws1,idx,idx%2==0)
    for col,w in zip("ABCDEFGHIJKL",[12,14,12,36,10,14,10,10,10,26,10,32]):
        ws1.column_dimensions[col].width=w
    ws2=wb.create_sheet("Reasoning Report"); ws2.append(["Date","Subject","Hall","Invigilator","Score","Why Selected","Duties"]); hr(ws2,1)
    for idx,r in enumerate(sorted(duty_records,key=sk),2):
        rea=r.get("reasoning",{})
        ws2.append([r.get("exam_date",""),r.get("subject",""),r.get("hall",""),r.get("faculty_name",""),
                    rea.get("score","")," | ".join(rea.get("why_selected",[])),rea.get("duty_count","")])
        dr(ws2,idx,idx%2==0)
    for col,w in zip("ABCDEFG",[14,34,10,26,8,60,8]): ws2.column_dimensions[col].width=w
    intel=_intelligence(); ws3=wb.create_sheet("Semester Intelligence")
    ws3["A1"]="Semester Intelligence Report"; ws3["A1"].font=Font(bold=True,size=13)
    ws3.append([]); ws3.append(["Metric","Value"]); hr(ws3,3)
    for i,(k,v) in enumerate([("Total duties",intel.get("total_duties",0)),("Exams covered",intel.get("total_exams",0)),
        ("Faculty with 0 duties",intel.get("zero_count",0)),("Mean duties",intel.get("mean_duties",0)),
        ("Busiest day",intel.get("busiest_day","N/A")),("Suggestion",intel.get("suggestion","N/A"))],4):
        ws3.append([k,str(v)]); dr(ws3,i,i%2==0)
    ws3.column_dimensions["A"].width=36; ws3.column_dimensions["B"].width=50
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  ENDPOINTS
# ══════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {"system":"BMS Invigilation v10","ready":timetable_done and exam_done,
            "faculty":len(faculty_map),"exams":len(exam_schedule),"duties":len(duty_records),
            "students_uploaded":students_done,"student_count":len(student_list),
            "multi_batch":multi_batch,"detected_usn_prefix":detected_usn_prefix,
            "rooms":_rooms_with_usn_ranges()}

@app.get("/status")
def status():
    return {"timetable_uploaded":timetable_done,"exam_uploaded":exam_done,
            "faculty_count":len(faculty_map),"exam_count":len(exam_schedule),
            "duty_count":len(duty_records),"students_uploaded":students_done,
            "student_count":len(student_list),"multi_batch":multi_batch,
            "detected_usn_prefix":detected_usn_prefix,"ready":timetable_done and exam_done}


# ── ROOMS ENDPOINT (NEW v10) ───────────────────────────────────
@app.get("/rooms")
def get_rooms():
    """Return current room configuration with live USN ranges."""
    return {"rooms": _rooms_with_usn_ranges(), "detected_prefix": detected_usn_prefix,
            "total_students": sum(r["students"] for r in room_config)}

@app.post("/rooms")
def update_rooms(payload: Dict):
    """
    Update room capacities. Send JSON:
    {"rooms": [{"room": "CA1", "students": 40}, {"room": "CA2", "students": 40}, ...]}
    Room names must match existing rooms. Capacities must be 1-200.
    Returns updated room config with recomputed USN ranges.
    """
    global room_config
    with state_lock:
        incoming = payload.get("rooms", [])
        if not incoming:
            raise HTTPException(400, "No rooms provided. Send {\"rooms\": [{\"room\": \"CA1\", \"students\": 40}, ...]}")
        new_config = []
        for r in incoming:
            name = str(r.get("room","")).strip().upper()
            try: cap = int(r.get("students", 0))
            except (ValueError, TypeError): raise HTTPException(400, f"Invalid student count for room {name}")
            if cap < 1 or cap > 200:
                raise HTTPException(400, f"Room {name}: student count must be 1–200, got {cap}")
            if not name:
                raise HTTPException(400, "Each room must have a 'room' name")
            new_config.append({"room":name,"students":cap,"usn_start":"","usn_end":"","usn_range":""})
        room_config = new_config
        save_rooms()
        return {"success":True,"message":f"{len(room_config)} room(s) updated.",
                "rooms":_rooms_with_usn_ranges(),"total_students":sum(r["students"] for r in room_config)}


@app.post("/upload_timetable")
async def upload_timetable(file: UploadFile = File(...)):
    global faculty_map, busy_slots, timetable_done, multi_batch
    fname = file.filename or "timetable.pdf"
    ext   = fname.lower().rsplit(".",1)[-1] if "." in fname else ""
    if ext not in ACCEPTED_EXTS: raise HTTPException(400, f"Unsupported: .{ext}. Use PDF, DOCX, XLSX, or CSV.")
    raw = await file.read()
    if not raw: raise HTTPException(400, "File is empty.")
    result = parse_timetable(raw, filename=fname)
    with state_lock:
        faculty_map.update(result["faculty_map"])
        busy_slots     = result["busy_slots"]
        multi_batch    = result.get("multi_batch", False)
        timetable_done = True
        return {"success":True,"file_type":ext.upper(),"faculty_detected":len(faculty_map),
                "faculty_map":faculty_map,"map_quality":result.get("map_quality","unknown"),
                "busy_slots":len(busy_slots),"multi_batch":multi_batch,
                "detected_codes":result.get("detected_codes",[]),"used_defaults":result.get("used_defaults",[]),
                "warnings":result.get("warnings",[]),"parse_quality":result.get("parse_quality","unknown"),
                "message":(f"Timetable parsed ({ext.upper()}). {len(faculty_map)} faculty, {len(busy_slots)} teaching slots."
                           + (" Multi-batch: clash detection ON." if multi_batch else " Single batch — clash detection OFF.")
                           + (f" Note: {result['used_defaults'][0]}" if result.get("used_defaults") else ""))}


@app.post("/upload_student_list")
async def upload_student_list(file: UploadFile = File(...)):
    """
    Upload student list (PDF/CSV/XLSX/DOCX).
    Columns needed: USN (any format), Student Name.
    Works for any year: 1BM25MC, 1BM26MC, 1BM23MCA, etc.
    After upload, attendance sheets auto-fill student names.
    """
    global student_list, students_done, detected_usn_prefix, room_config
    fname = file.filename or "students.pdf"
    ext   = fname.lower().rsplit(".",1)[-1] if "." in fname else ""
    if ext not in ACCEPTED_EXTS: raise HTTPException(400, f"Unsupported: .{ext}. Use PDF, CSV, XLSX, or DOCX.")
    raw = await file.read()
    if not raw: raise HTTPException(400, "File is empty.")
    result   = parse_student_list(raw, filename=fname)
    students = result.get("students", [])
    if result.get("total", 0) == 0 or not students:
        raise HTTPException(
            status_code=400,
            detail="No students parsed. Check file format."
        )



    with state_lock:
        # KEY FIX: always store with normalise_usn key
        student_list = {
        normalise_usn(s["usn"]): {
            "name": s["name"],
            "programme": s.get("programme", "MCA")
        }
        for s in students
    }
        students_done = bool(student_list)

        # Update detected prefix and rebuild room USN ranges from actual student data
        prefix = result.get("detected_prefix", "")
        if prefix:
            detected_usn_prefix = prefix
            room_config = rebuild_rooms_with_prefix(_current_rooms(), prefix, len(student_list))
            save_rooms()

        return {"success":True,"students_loaded":len(student_list),
                "parse_quality":result.get("parse_quality","unknown"),
                "detected_prefix":detected_usn_prefix,
                "warnings":result.get("warnings",[]),
                "rooms_updated":_rooms_with_usn_ranges(),
                "sample":list(student_list.items())[:5],
                "message":(f"{len(student_list)} student(s) loaded from {fname}. "
                           + (f"USN prefix detected: {detected_usn_prefix}. " if detected_usn_prefix else "")
                           + "Attendance sheets will auto-fill names.")}

    


@app.post("/upload_exam_and_allocate")
async def upload_exam_and_allocate(file: UploadFile = File(...)):
    global exam_schedule, duty_records, exam_done, last_result
    if not timetable_done: raise HTTPException(400, "Upload timetable first (Step 1).")
    fname = file.filename or "exam.pdf"
    ext   = fname.lower().rsplit(".",1)[-1] if "." in fname else ""
    if ext not in ACCEPTED_EXTS: raise HTTPException(400, f"Unsupported: .{ext}.")
    raw = await file.read()
    if not raw: raise HTTPException(400, "File is empty.")
    with state_lock:
        # Pass current room config so exam schedule uses correct rooms
        parsed        = parse_exam_schedule(raw, filename=fname, custom_rooms=_rooms_with_usn_ranges())
        exam_schedule = parsed["exams"]
        exam_done     = True
        duty_records  = []   # always reset before re-allocating
        allocation    = _run_all()
        last_result   = allocation
        return {"success":True,"file_type":ext.upper(),"variant":parsed.get("variant",""),
                "exams_found":parsed["total"],"rooms_detected":parsed.get("rooms",[]),
                "used_defaults":parsed.get("used_defaults",[]),"warnings":parsed.get("warnings",[]),
                "parse_quality":parsed.get("parse_quality","unknown"),
                "allocation":allocation,"intelligence":_intelligence(),
                "message":(f"Exam file parsed ({ext.upper()}). {parsed['total']} exam(s) loaded. "+allocation["message"])}


@app.get("/results")
def get_results():
    if not last_result: raise HTTPException(400, "Run allocation first.")
    return {**last_result, "intelligence": _intelligence()}

@app.get("/clashes")
def clashes():
    if not multi_batch:
        return {"total":0,"clashes":[],"note":"Single-batch timetable — no clashes possible. "
                "Clashes activate once junior batches join and combined timetable is uploaded."}
    found = []
    for exam in exam_schedule:
        for code, name in faculty_map.items():
            if _is_teaching(code, exam.get("day",""), exam.get("start_time","")):
                found.append({"faculty":name,"code":code,"exam_subject":exam.get("subject",""),
                               "exam_date":exam.get("exam_date",""),"exam_time":exam.get("start_time",""),
                               "day":exam.get("day",""),"action":"Auto-excluded from this exam"})
    return {"total":len(found),"clashes":found}

@app.get("/workload")
def workload():
    counts = {n: _duty_count(n) for n in faculty_map.values() if n}
    vals   = list(counts.values()); mean = round(statistics.mean(vals),1) if vals else 0
    return {"faculty":sorted([{"name":n,"duties":c} for n,c in counts.items()],key=lambda x:-x["duties"]),
            "total_duties":len(duty_records),"mean_duties":mean}

@app.get("/semester_intelligence")
def semester_intelligence(): return _intelligence()

@app.get("/export_duty_chart")
def export_duty_chart():
    if not duty_records: raise HTTPException(400, "Run allocation first.")
    return Response(content=_build_duty_excel(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":"attachment; filename=bms_duty_chart.xlsx"})

@app.get("/export_attendance_sheets")
def export_attendance_sheets():
    if not last_result: raise HTTPException(400, "Run allocation first.")
    sheets = generate_attendance_sheets(last_result.get("results",[]))
    return Response(content=sheets,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":"attachment; filename=bms_attendance_sheets.xlsx"})

@app.get("/duty_history")
def duty_history():
    return {"total":len(duty_records),
            "records":sorted([{k:v for k,v in r.items() if k!="reasoning"} for r in duty_records],
                             key=lambda r:(parse_date(r.get("exam_date","")) or datetime.min,
                                           parse_time(r.get("start_time","")) or datetime.min))}

@app.delete("/reset")
def reset(what: str = Query(default="duties")):
    global faculty_map,busy_slots,exam_schedule,duty_records,timetable_done,exam_done,last_result
    global student_list,students_done,multi_batch,room_config,detected_usn_prefix
    with state_lock:
        if what == "all":
            faculty_map={}; busy_slots=[]; exam_schedule=[]; duty_records=[]
            student_list={}; students_done=False; multi_batch=False
            room_config=deepcopy(DEFAULT_ROOMS); detected_usn_prefix=""
            timetable_done=False; exam_done=False; last_result={}
            save_rooms()
            return {"message":"All data cleared. Room config reset to defaults."}
        n=len(duty_records); duty_records=[]; last_result={}
        return {"message":f"{n} duty record(s) cleared. Room config preserved."}


def load_rooms():
    global room_config
    try:
        with open(ROOM_FILE, "r") as f:
            room_config = json.load(f)
            log_info("Rooms loaded from file")
    except Exception:
        room_config = deepcopy(DEFAULT_ROOMS)
        log_info("Using default rooms")


def save_rooms():
    try:
        with open(ROOM_FILE, "w") as f:
            json.dump(room_config, f, indent=2)
        log_info("Rooms saved")
    except Exception as e:
        log_error(f"Room save failed: {e}")
load_rooms()