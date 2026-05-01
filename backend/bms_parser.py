"""
BMS College — Universal File Parser (v9 — All Bugs Fixed)
==========================================================
BUGS FIXED vs v8:
  BUG 1: Short day abbreviations (Mon/MON/Tue/TUE) now recognised
  BUG 2: 24-hour time (14:00, 09:15) now parsed correctly
  BUG 3: Grid-format timetable (day=column header) now correctly parsed
  BUG 4: Multi-batch detection regex fixed for Roman numerals I/II/III
  BUG 5: DOCX table handles both row-per-day AND col-per-day layouts
  BUG 6: Subject regex protected from matching time strings and room codes
  BUG 7: _SLOT_RE made non-greedy, prevents runaway matches
  BUG 8: Fallback strategy added when no day headers found at all

Accepts: PDF (digital/scanned), DOCX, XLSX, XLS, CSV
"""
import re, io, csv as csvlib, os
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime, timedelta
from copy import deepcopy
from utils import normalise_usn, log_error, log_info

DEFAULT_STUDENTS = 120
DEFAULT_ROOMS = [
    {"room":"CA1","students":30,"usn_start":"1BM25MCA001","usn_end":"1BM25MCA030","usn_range":"1BM25MCA001-030"},
    {"room":"CA2","students":30,"usn_start":"1BM25MCA031","usn_end":"1BM25MCA060","usn_range":"1BM25MCA031-060"},
    {"room":"BT1","students":30,"usn_start":"1BM25MCA061","usn_end":"1BM25MCA090","usn_range":"1BM25MCA061-090"},
    {"room":"BT2","students":26,"usn_start":"1BM25MCA091","usn_end":"1BM25MCA116","usn_range":"1BM25MCA091-116"},
]
KNOWN_FACULTY = {
    "SU":"Dr. S. Uma","DNS":"Dr. D. N. Sujatha","RMR":"Dr. Ch. Ram Mohan Reddy",
    "VK":"Dr. K. Vijaya Kumar","VPP":"Dr. V. Padmapriya","TSP":"Smt. Pushpa T.S.",
    "GK":"Sri. Girish K.","KPS":"Smt. K.P. Shailaja","RR":"Sri. R.V. Raghavendra Rao",
    "SS":"Smt. S. Shilpa","TS":"Smt. T. Sunitha","NM":"Sri. Nagaraj M.K.",
    "VR":"Veena R.","NF":"New Faculty",
}
KNOWN_SUBJECTS = [
    "Mathematical Foundations for Computer Applications",
    "Database Management Systems","Web Technologies","Python Programming",
    "Java Programming","Machine Learning","Deep Learning","Computer Networks",
    "Data Structures using C","DS using C","Software Project Management",
    "Unix Laboratory","Design and Analysis of Algorithms","Operating Systems",
    "Discrete Mathematics","Computer Organization","Advanced Java","Data Science",
    "Cloud Computing","Artificial Intelligence",
]
HARD_ROOMS = {
    "CA1","CA2","CA3","BT1","BT2","LAB1A","LAB1B","LAB2","Lab1A","Lab1B",
    "RL","FDC","DLA","DLB","DL1","DL2","DL3","DL4","DL5","DL6","NET1","NET2",
}
TESSERACT_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    "/usr/bin/tesseract","/usr/local/bin/tesseract",
]
NAME_PREFIXES = ("Dr.","Smt.","Sri.","Prof.","Mr.","Ms.")

# FIX 1 — complete day lookup including abbreviations
DAY_MAP = {
    "MONDAY":"Monday","TUESDAY":"Tuesday","WEDNESDAY":"Wednesday",
    "THURSDAY":"Thursday","FRIDAY":"Friday","SATURDAY":"Saturday",
    "MON":"Monday","TUE":"Tuesday","WED":"Wednesday",
    "THU":"Thursday","THUR":"Thursday","THURS":"Thursday",
    "FRI":"Friday","SAT":"Saturday",
}
_DAY_KEYS = sorted(DAY_MAP.keys(), key=len, reverse=True)
_DAY_PATTERN = "|".join(_DAY_KEYS)

# Codes / words that must never be treated as faculty codes
_BAD_CODES = {"AM","PM","TO","THE","AND","FOR","OF","IN","ON","AT","BY","IS","IT"}

def generate_usns(usn_start: str, usn_end: str) -> List[str]:
    try:
        ms=re.search(r'(\d+)$',usn_start); me=re.search(r'(\d+)$',usn_end)
        if not (ms and me): return [usn_start,usn_end]
        prefix=usn_start[:ms.start()]; nlen=len(ms.group(1))
        sn,en=int(ms.group(1)),int(me.group(1))
        return [f"{prefix}{str(i).zfill(nlen)}" for i in range(sn,en+1)]
    except: return [usn_start]

# ── Date / Time ────────────────────────────────────────────────
DATE_FMTS=["%d/%m/%Y","%d-%m-%Y","%d-%b-%Y","%d %b %Y","%d/%m/%y","%d-%m-%y","%d %B %Y"]

def parse_date(s:Any)->Optional[datetime]:
    if not s: return None
    s=str(s).strip(); s=re.sub(r"\([^)]*\)","",s)
    s=re.sub(r"\b(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b","",s,flags=re.IGNORECASE).strip().strip(",").strip()
    for fmt in DATE_FMTS:
        try: return datetime.strptime(s,fmt)
        except: pass
    return None

def parse_time(s:Any)->Optional[datetime]:
    if not s: return None
    s=str(s).strip().replace("\u00a0"," ").replace(" ","").upper()
    s=re.sub(r"^(\d{1,2})(\d{2})(AM|PM)$",r"\1.\2\3",s)
    for fmt in ["%I.%M%p","%I:%M%p","%H.%M","%H:%M","%I%p","%I.%M %p","%I:%M %p"]:
        try: return datetime.strptime(s,fmt)
        except: pass
    for fmt in ["%I:%M%p","%H:%M"]:
        try: return datetime.strptime(s.replace(".",":",),fmt)
        except: pass
    return None

def fmt_time(dt): return dt.strftime("%I:%M %p").lstrip("0") if dt else ""
def fmt_date(dt): return dt.strftime("%d-%b-%Y") if dt else ""

# FIX 2 — find_time_range handles 24h times
def find_time_range(text:str)->Tuple[str,str]:
    if not text: return "",""
    SEP = r"\s*(?:[Tt][oO]|–|—|-|TO)\s*"
    # With AM/PM
    p1=re.compile(
        r"(\d{1,2}[.:]\d{2}\s*(?:AM|PM|am|pm)|\d{1,2}\s*(?:AM|PM|am|pm))"
        + SEP +
        r"(\d{1,2}[.:]\d{2}\s*(?:AM|PM|am|pm)|\d{1,2}\s*(?:AM|PM|am|pm))",
        re.IGNORECASE)
    m=p1.search(text)
    if m:
        t1,t2=fmt_time(parse_time(m.group(1))),fmt_time(parse_time(m.group(2)))
        if t1 and t2: return t1,t2
    # 24-hour without AM/PM
    p2=re.compile(r"(\d{1,2}[.:]\d{2})" + SEP + r"(\d{1,2}[.:]\d{2})", re.IGNORECASE)
    m2=p2.search(text)
    if m2:
        t1,t2=fmt_time(parse_time(m2.group(1))),fmt_time(parse_time(m2.group(2)))
        if t1 and t2: return t1,t2
    return "",""

# ── PDF extraction ─────────────────────────────────────────────
def _page_useful(t):
    t=t.strip(); return len(t)>=50 and sum(1 for c in t if c.islower())>=15

def _setup_tess():
    try:
        import pytesseract
        for p in TESSERACT_PATHS:
            if os.path.exists(p): pytesseract.pytesseract.tesseract_cmd=p; return True
        pytesseract.get_tesseract_version(); return True
    except: return False

def _ocr(img):
    try:
        import pytesseract; _setup_tess()
        return pytesseract.image_to_string(img,config="--oem 3 --psm 6",lang="eng")
    except: return ""

def extract_text_from_pdf(fb:bytes)->Tuple[str,List[str]]:
    warns=[]
    try: import pdfplumber
    except ImportError: return "",["pdfplumber not installed"]
    pages,scanned=[],[]
    try:
        with pdfplumber.open(io.BytesIO(fb)) as pdf:
            for i,page in enumerate(pdf.pages,1):
                try:
                    t=(page.extract_text(x_tolerance=4,y_tolerance=4) or "").strip()
                    if _page_useful(t): pages.append(t)
                    else: scanned.append(i); pages.append(f"<<OCR:{i}>>")
                except: pages.append("")
    except Exception as e: return "",["PDF read error: "+str(e)[:80]]
    if scanned:
        ok=False
        try: import pdf2image; ok=_setup_tess()
        except: pass
        if ok:
            warns.append(f"Pages {scanned} scanned — running OCR (~{len(scanned)*15}s)")
            try:
                from pdf2image import convert_from_bytes
                imgs=convert_from_bytes(fb,dpi=250)
                for pn in scanned:
                    if pn-1<len(imgs):
                        txt=_ocr(imgs[pn-1])
                        for j,p in enumerate(pages):
                            if p==f"<<OCR:{pn}>>": pages[j]=txt; break
            except Exception as e: warns.append("OCR failed: "+str(e)[:60])
        else: warns.append("Scanned PDF — install pytesseract+pdf2image for OCR")
    full="\n\n".join(p for p in pages if p and not p.startswith("<<OCR:"))
    return full,warns

def extract_text_from_docx(fb:bytes)->Tuple[str,List[List[List[str]]],List[str]]:
    warns=[]
    try: from docx import Document
    except: return "",[],["python-docx not installed"]
    try:
        doc=Document(io.BytesIO(fb))
        txt="\n".join(p.text for p in doc.paragraphs if p.text.strip())
        tables=[]
        for table in doc.tables:
            rows=[]
            for row in table.rows:
                cells=[cell.text.strip().replace('\n',' ') for cell in row.cells]
                rows.append(cells)
            tables.append(rows)
            for row in table.rows:
                rt=" | ".join(c.text.strip() for c in row.cells if c.text.strip())
                if rt: txt+="\n"+rt
        return txt,tables,warns
    except Exception as e: return "",[],["DOCX error: "+str(e)[:100]]

def extract_text_from_xlsx(fb:bytes)->Tuple[str,List[List],List[str]]:
    warns,rows_data=[],[]
    try:
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(fb),read_only=True)
        ws=wb.active; full=""
        for row in ws.iter_rows(values_only=True):
            cells=[str(c).strip() if c is not None else "" for c in row]
            if any(cells): rows_data.append(cells); full+=" | ".join(c for c in cells if c)+"\n"
        return full,rows_data,warns
    except:
        try:
            import pandas as pd; df=pd.read_excel(io.BytesIO(fb))
            return df.to_string(),df.values.tolist(),warns
        except Exception as e: return "",[],["XLSX error: "+str(e)[:100]]

# ── Faculty extraction ──────────────────────────────────────────
def extract_faculty_from_docx_tables(tables):
    mapping={}
    for table in tables:
        for row in table:
            name_cell=code_cell=""
            for cell in row:
                cell=cell.strip()
                if any(cell.startswith(p) for p in NAME_PREFIXES) and len(cell)>4: name_cell=cell
                elif re.match(r'^[A-Z]{2,5}$',cell) and cell not in HARD_ROOMS: code_cell=cell
            if name_cell and code_cell: mapping[code_cell]=name_cell
    return mapping

def extract_faculty_from_text(text:str)->Dict[str,str]:
    mapping={}; lines=[l.strip() for l in text.splitlines() if l.strip()]
    p1=re.compile(r"(?:\d+[.)]\s*)?((?:Dr\.|Smt\.|Sri\.|Prof\.|Mr\.|Ms\.)\s+[A-Za-z.\s]+?)\s{2,}([A-Z]{2,5})\s*$",re.MULTILINE)
    for m in p1.finditer(text):
        code=m.group(2).strip()
        if code not in HARD_ROOMS: mapping[code]=m.group(1).strip().rstrip(".")
    for i in range(len(lines)-1):
        l0,l1=lines[i],lines[i+1]
        if any(l0.startswith(p) for p in NAME_PREFIXES):
            if re.match(r'^([A-Z]{2,5})$',l1) and l1 not in HARD_ROOMS: mapping.setdefault(l1,l0)
    p3=re.compile(r"\d+[.)]\s+((?:Dr\.|Smt\.|Sri\.|Prof\.|Mr\.|Ms\.)\s+[A-Za-z.\s]+?)\s+([A-Z]{2,5})\b")
    for m in p3.finditer(text):
        code=m.group(2).strip()
        if code not in HARD_ROOMS: mapping.setdefault(code,m.group(1).strip())
    p4=re.compile(r"\|?\s*((?:Dr\.|Smt\.|Sri\.|Prof\.)\s+[A-Za-z.\s]+?)\s*\|\s*([A-Z]{2,5})\s*\|?")
    for m in p4.finditer(text):
        code=m.group(2).strip()
        if code not in HARD_ROOMS: mapping.setdefault(code,m.group(1).strip())
    return mapping

def get_faculty_map(text:str,tables=None)->Tuple[Dict,str,List[str]]:
    defs=[]
    if tables:
        ex=extract_faculty_from_docx_tables(tables)
        if ex: return ex,"extracted_docx",[]
    ex=extract_faculty_from_text(text)
    if ex: return ex,"extracted_text",[]
    found=set(re.findall(r'\b([A-Z]{2,5})\b',text))
    known={c:n for c,n in KNOWN_FACULTY.items() if c in found}
    if known: defs.append(f"Used built-in BMS codes — {len(known)} matched"); return known,"known_codes",defs
    defs.append("No faculty found — using full built-in BMS list"); return dict(KNOWN_FACULTY),"default",defs

# ══════════════════════════════════════════════════════════════
# BUSY SLOT EXTRACTION — fully rewritten (FIX 1,3,5,6,7,8)
# ══════════════════════════════════════════════════════════════

# Core slot regex: SUBJECT (CODE1, CODE2)
# Protected: won't match pure time strings or room codes
_SLOT_RE = re.compile(
    r"([A-Za-z][A-Za-z\s&/\-]{1,40}?)"   # subject (lazy, min 1 letter)
    r"\s*\(\s*"
    r"([A-Z]{2,5}(?:\s*,\s*[A-Z]{2,5})*)"  # codes
    r"\s*\)",
)

def _clean_subj(raw:str)->str:
    s=re.sub(r'^\s*[\d:.\-\s]+','',raw)
    s=re.sub(r'\s{2,}',' ',s).strip().strip("|(").strip()
    if len(s)<2: return ""
    if re.search(r'\d{1,2}[:.]\d{2}',s): return ""  # looks like a time
    if s.upper() in DAY_MAP: return ""               # is a day name
    if s.lower() in {"am","pm","to","st","nd","rd","th","lab","room","hall",
                     "batch","section","sem","semester","year","class","break",
                     "lunch","interval","period"}: return ""
    return s[:60]

def _valid_code(c:str)->bool:
    return (len(c)>=2 and c not in HARD_ROOMS and not c.isdigit()
            and c not in _BAD_CODES)

def _dedup(slots:List[Dict])->List[Dict]:
    seen,out=set(),[]
    for s in slots:
        k=(s["faculty_code"],s["day"])
        if k not in seen: seen.add(k); out.append(s)
    return out

def _slot(code,subj,day): return {"faculty_code":code,"subject":subj,"day":day,"time_slot":"","room":""}

# FIX 1+3+8: handles full names, abbreviations, grid format, inline format
def extract_busy_slots_from_text(text:str)->List[Dict]:
    if not text: return []
    slots=[]

    # Strategy A: linear text — day header lines then slots below
    day_line_re=re.compile(r"^\s*("+_DAY_PATTERN+r")\s*$",re.IGNORECASE|re.MULTILINE)
    positions=[]
    for m in day_line_re.finditer(text):
        raw=m.group(1).upper().strip()
        positions.append((m.start(), DAY_MAP.get(raw, raw.capitalize())))
    positions.sort(key=lambda x:x[0])

    if positions:
        for i,(start,day) in enumerate(positions):
            end=positions[i+1][0] if i+1<len(positions) else len(text)
            section=text[start:end]
            for m in _SLOT_RE.finditer(section):
                subj=_clean_subj(m.group(1)); 
                if not subj: continue
                for raw in re.split(r'\s*,\s*',m.group(2)):
                    code=raw.strip().upper()
                    if _valid_code(code): slots.append(_slot(code,subj,day))

    # Strategy B: grid — multiple day names on one line (column headers)
    grid_re=re.compile(r"(?:"+_DAY_PATTERN+r")(?:\s+|\s*\|\s*)(?:"+_DAY_PATTERN+r")",re.IGNORECASE)
    if grid_re.search(text) and not slots:
        slots.extend(_parse_grid_text(text))

    # Strategy C: inline "Day: Subject (CODE)"
    if not slots:
        inline=re.compile(
            r"("+_DAY_PATTERN+r")\s*[:\-]\s*([A-Za-z][A-Za-z\s&/\-]{1,40}?)"
            r"\s*\(\s*([A-Z]{2,5}(?:\s*,\s*[A-Z]{2,5})*)\s*\)",re.IGNORECASE)
        for m in inline.finditer(text):
            day=DAY_MAP.get(m.group(1).upper(),m.group(1).capitalize())
            subj=_clean_subj(m.group(2))
            if not subj: continue
            for raw in re.split(r'\s*,\s*',m.group(3)):
                code=raw.strip().upper()
                if _valid_code(code): slots.append(_slot(code,subj,day))

    # Strategy D: absolute fallback — any (CODE) pattern, day=Unknown
    if not slots:
        for m in _SLOT_RE.finditer(text):
            subj=_clean_subj(m.group(1))
            if not subj: continue
            for raw in re.split(r'\s*,\s*',m.group(2)):
                code=raw.strip().upper()
                if _valid_code(code): slots.append(_slot(code,subj,"Unknown"))

    if not slots:
        log_error("No timetable slots detected")

    # DEBUG VIEW
        preview = text[:1000]
        log_info("TEXT PREVIEW:")
        log_info(preview)

    # Try fallback simple extraction
        simple_matches = re.findall(r'[A-Z]{2,5}\d{2,3}', text)

        log_info(f"Fallback found {len(simple_matches)} codes")


    if not slots:
        log_error("No timetable slots detected. Showing preview:")
        log_info(text[:1000])      
                

    return _dedup(slots)

def _parse_grid_text(text:str)->List[Dict]:
    """Parse grid where days are column headers."""
    slots=[]; lines=[l.strip() for l in text.splitlines() if l.strip()]
    if not lines: return slots
    day_re=re.compile(r'\b('+_DAY_PATTERN+r')\b',re.IGNORECASE)
    header_idx=-1; header_days=[]; header_line=""
    for i,line in enumerate(lines):
        found=day_re.findall(line)
        if len(found)>=2:
            header_idx=i; header_line=line
            header_days=[DAY_MAP.get(d.upper(),d.capitalize()) for d in found]
            break
    if header_idx==-1: return slots
    col_positions=[]
    for d in header_days:
        m=re.search(re.escape(d),header_line,re.IGNORECASE)
        if m: col_positions.append(m.start())
    if not col_positions: return slots
    for line in lines[header_idx+1:]:
        if day_re.search(line) and len(day_re.findall(line))>=2: continue
        for m in _SLOT_RE.finditer(line):
            pos=m.start()
            day=header_days[0]
            for ci,cp in enumerate(col_positions):
                if cp<=pos: day=header_days[ci]
            subj=_clean_subj(m.group(1))
            if not subj: continue
            for raw in re.split(r'\s*,\s*',m.group(2)):
                code=raw.strip().upper()
                if _valid_code(code): slots.append(_slot(code,subj,day))
    return slots

# FIX 5+6: DOCX handles both table orientations
def extract_busy_slots_from_docx(tables:List[List[List[str]]])->List[Dict]:
    slots=[]
    if not tables: return slots
    for table in tables:
        if not table or len(table)<2: continue
        # Detect orientation by counting day names in first row vs first column
        first_row=[c.strip().upper() for c in table[0]]
        first_col=[row[0].strip().upper() if row else "" for row in table]
        days_in_row=sum(1 for c in first_row if c in DAY_MAP or c[:3] in DAY_MAP)
        days_in_col=sum(1 for c in first_col if c in DAY_MAP or c[:3] in DAY_MAP)
        if days_in_row>=2: slots.extend(_docx_col_days(table))
        else:              slots.extend(_docx_row_days(table))
    return _dedup(slots)

def _docx_row_days(table:List[List[str]])->List[Dict]:
    slots=[]; cur_day="Unknown"
    for row in table:
        if not row: continue
        first=row[0].strip().upper()
        day=DAY_MAP.get(first) or DAY_MAP.get(first[:3])
        if day: cur_day=day
        for cell in row:
            if not cell: continue
            for m in _SLOT_RE.finditer(cell):
                subj=_clean_subj(m.group(1))
                if not subj: continue
                for raw in re.split(r'\s*,\s*',m.group(2)):
                    code=raw.strip().upper()
                    if _valid_code(code): slots.append(_slot(code,subj,cur_day))
    return slots

def _docx_col_days(table:List[List[str]])->List[Dict]:
    slots=[]; header=table[0]; col_day={}
    for j,cell in enumerate(header):
        raw=cell.strip().upper()
        day=DAY_MAP.get(raw) or DAY_MAP.get(raw[:3])
        if day: col_day[j]=day
    if not col_day: return slots
    for row in table[1:]:
        for j,cell in enumerate(row):
            if j not in col_day or not cell: continue
            day=col_day[j]
            for m in _SLOT_RE.finditer(cell):
                subj=_clean_subj(m.group(1))
                if not subj: continue
                for raw in re.split(r'\s*,\s*',m.group(2)):
                    code=raw.strip().upper()
                    if _valid_code(code): slots.append(_slot(code,subj,day))
    return slots

# FIX 4: multi-batch detection with Roman numerals
def detect_multiple_batches(text:str)->bool:
    patterns=[
        r'\b(?:I{1,3}V?|IV|V?I{0,3})\s*(?:Sem|Semester|SEM)\b',  # Roman: I Sem, II Sem
        r'\b\d\s*(?:st|nd|rd|th)\s*(?:Sem|Semester)\b',            # 1st Sem, 2nd Sem
        r'\bBatch\s*[A-D1-4]\b',
        r'\bSection\s*[A-D]\b',
        r'\b(?:1st|2nd|3rd|4th|First|Second|Third|Fourth)\s*Year\b',
        r'\b(?:Junior|Senior)\s*(?:Batch|Class|Students)\b',
    ]
    # Must match at least 2 distinct patterns to confirm multi-batch
    matched=sum(1 for p in patterns if re.search(p,text,re.IGNORECASE))
    return matched>=2

# ── Room extraction ────────────────────────────────────────────
def extract_rooms(text:str)->List[Dict]:
    rooms,seen=[],set()
    for m in re.compile(r'\b(CA\s*\d|BT\s*\d)\b(.*?)\b(\d{2,3})\s*(?:\n|$)',re.IGNORECASE|re.MULTILINE).finditer(text):
        room=re.sub(r'\s+','',m.group(1)).upper(); count=int(m.group(3)); rest=m.group(2).strip()
        if 5<=count<=200 and room not in seen:
            seen.add(room); usns=re.findall(r'(1\s*BM\s*\d{2}\s*MC?\s*\d{3})',rest,re.IGNORECASE)
            rooms.append({"room":room,"students":count,"usn_range":rest[:60],
                          "usn_start":usns[0] if len(usns)>=1 else "",
                          "usn_end":usns[-1] if len(usns)>=2 else (usns[0] if usns else "")})
    if not rooms:
        for m in re.finditer(r'\b(CA\d|BT\d)\b[^\n]{0,50}?(\d{2,3})\b',text,re.IGNORECASE):
            room,count=m.group(1).upper(),int(m.group(2))
            if 5<=count<=200 and room not in seen:
                seen.add(room); rooms.append({"room":room,"students":count,"usn_range":"","usn_start":"","usn_end":""})
    return rooms if rooms else deepcopy(DEFAULT_ROOMS)

def detect_subject(text_block:str)->str:
    combined=re.sub(r'\s{2,}',' ',text_block.strip()); low=combined.lower()
    for s in KNOWN_SUBJECTS:
        if s.lower() in low: return s
    for s in KNOWN_SUBJECTS:
        words=[w for w in s.split() if len(w)>4]
        if len(words)>=2 and sum(1 for w in words if w.lower() in low)>=len(words)-1: return s
    skip={"seating","room","usn","date","time","note","head","controller","principal",
          "coordinator","bms","college","engineering","vtu","department"}
    for line in combined.splitlines():
        line=line.strip()
        if (len(line)>12 and re.search(r'[a-zA-Z]{4,}',line)
                and not re.search(r'\d{2}[/\-]\d{2}',line)
                and not re.match(r'^[A-Z]{2,5}\s*$',line)
                and not any(k in line.lower() for k in skip)):
            return line[:80]
    return ""

# ── Student list parser ────────────────────────────────────────
def parse_student_list(file_bytes:bytes,filename:str="")->Dict:
    if not file_bytes:
        return {"students":[],"total":0,"warnings":["Empty file"],"parse_quality":"fallback"}
    ext=filename.lower().rsplit(".",1)[-1] if "." in filename else "csv"
    if ext=="csv": return _students_csv(file_bytes)
    if ext in("xlsx","xls","ods"): return _students_xlsx(file_bytes)
    if ext in("docx","doc"): return _students_docx(file_bytes)
    if ext=="pdf": return _students_pdf(file_bytes)
    return {"students":[],"total":0,"warnings":[f"Unsupported: {ext}"],"parse_quality":"fallback"}

def _normalise_student_row(row_dict:Dict)->Optional[Dict]:
    n={str(k).strip().lower().replace(" ","_"):str(v or "").strip() for k,v in row_dict.items() if k is not None}
    usn=""
    for key in("usn","roll_no","roll","enrollment","enrolment","reg_no","registration","usn/roll"):
        if key in n and n[key]: usn=normalise_usn(n[key]); break
    if not usn:
        for k,v in n.items():
            if re.match(r'(1\s*BM\s*\d{2}\s*MC?\s*\d{3})',v,re.IGNORECASE): usn=normalise_usn(v); break
    name=""
    for key in("name","student_name","full_name","student","candidate_name"):
        if key in n and n[key]: name=n[key].strip().title(); break
    if not usn or not name: return None
    prog="MCA"
    for key in("programme","program","course","branch","dept"):
        if key in n and n[key]: prog=n[key].strip(); break
    return {"usn":usn,"name":name,"programme":prog}

def _students_csv(fb):
    warns=[]
    try:
        text=fb.decode("utf-8-sig"); reader=csvlib.DictReader(io.StringIO(text)); students=[]
        for i,row in enumerate(reader,1):
            s=_normalise_student_row(row)
            if s: students.append(s)
            else: warns.append(f"Row {i}: could not extract USN/name")
        return {"students":students,"total":len(students),"warnings":warns,
                "parse_quality":"good" if students else "fallback"}
    except Exception as e:
        return {"students":[],"total":0,"warnings":[str(e)],"parse_quality":"fallback"}

def _students_xlsx(fb):
    warns=[]
    try:
        import openpyxl
        wb=openpyxl.load_workbook(io.BytesIO(fb),read_only=True)
        ws=wb.active; rows=list(ws.iter_rows(values_only=True))
        if not rows: return {"students":[],"total":0,"warnings":["Empty sheet"],"parse_quality":"fallback"}
        header=[str(c or "").strip().lower() for c in rows[0]]; students=[]
        for i,row in enumerate(rows[1:],2):
            row_dict={header[j]:str(row[j] or "").strip() for j in range(min(len(header),len(row)))}
            s=_normalise_student_row(row_dict)
            if s: students.append(s)
            else: warns.append(f"Row {i}: could not extract USN/name")
        return {"students":students,"total":len(students),"warnings":warns,
                "parse_quality":"good" if students else "fallback"}
    except Exception as e:
        return {"students":[],"total":0,"warnings":[str(e)],"parse_quality":"fallback"}

def _students_docx(fb):
    text,tables,warns=extract_text_from_docx(fb); students=[]
    if tables:
        for table in tables:
            if not table: continue
            header=[c.strip().lower() for c in table[0]]
            for row in table[1:]:
                row_dict={header[j]:row[j] for j in range(min(len(header),len(row)))}
                s=_normalise_student_row(row_dict)
                if s: students.append(s)
    if not students:
        for m in re.finditer(r'(1BM\d+MCA\d+)\s+([A-Z][a-zA-Z\s.]+)',text):
            students.append({"usn":m.group(1).upper(),"name":m.group(2).strip().title(),"programme":"MCA"})
    return {"students":students,"total":len(students),"warnings":warns,
            "parse_quality":"good" if students else "partial"}

def _students_pdf(fb):
    text, warns = extract_text_from_pdf(fb)
    students = []

    lines = text.splitlines()

    buffer = ""

    for raw_line in lines:
        line = raw_line.strip()

        if not line:
            continue

        # 🔥 Merge broken lines (important)
        if re.match(r'^\d+\s+1BM\d{2}MC\d{3}', line):
            if buffer:
                _process_line(buffer, students)
            buffer = line
        else:
            buffer += " " + line

    if buffer:
        _process_line(buffer, students)

    unique_students = list({s['usn']: s for s in students}.values())

    return {
        "students": unique_students,
        "total": len(unique_students),
        "warnings": warns,
        "parse_quality": "high" if unique_students else "failed"
    }


def _process_line(line, students):
    parts = line.split()

    if len(parts) < 5:
        return

    # Remove serial number
    if parts[0].isdigit():
        parts = parts[1:]

    usn = parts[0]

    # Strict validation
    if not re.match(r'1BM\d{2}MC\d{3}', usn):
        return

    # Remove USN
    parts = parts[1:]

    # Remove last numeric (resume ID)
    if parts and parts[-1].isdigit():
        parts = parts[:-1]

    # Now remaining:
    # [student name..., father..., mother...]

    if len(parts) >= 4:
        name_parts = parts[:-2]
    else:
        name_parts = parts

    name = " ".join(name_parts).title()

    if name:
        students.append({
            "usn": usn,
            "name": name,
            "programme": "MCA"
        })


def _process_robust_line(line, students):
    usn_match = re.search(r'1BM\d{2}MC\d{3}', line)
    if not usn_match:
        return
        
    usn = usn_match.group(0)

    content_after_usn = line[usn_match.end():].strip()

    # remove trailing numbers
    content_after_usn = re.sub(r'\s+\d+$', '', content_after_usn).strip()

    parts = content_after_usn.split()

    # remove trailing numeric again if needed
    if parts and parts[-1].isdigit():
        parts = parts[:-1]

    # safer name logic
    if len(parts) >= 5:
        name_parts = parts[:-2]
    else:
        name_parts = parts

    name = " ".join(name_parts).title()

    if name:
        students.append({
            "usn": usn,
            "name": name,
            "programme": "MCA"
        })
# ── Master parsers ─────────────────────────────────────────────
def parse_timetable(file_bytes:bytes,filename:str="")->Dict:
    if not file_bytes:
        return {"faculty_map":dict(KNOWN_FACULTY),"busy_slots":[],"detected_codes":[],
                "total_slots":0,"warnings":[],"used_defaults":["Empty file — built-in BMS list used"],
                "parse_quality":"fallback","multi_batch":False}
    ext=filename.lower().rsplit(".",1)[-1] if "." in filename else "pdf"
    if ext=="csv": return _tt_csv(file_bytes)
    if ext in("docx","doc"): return _tt_docx(file_bytes)
    if ext in("xlsx","xls","ods"): return _tt_xlsx(file_bytes)
    return _tt_pdf(file_bytes)

def _tt_docx(fb):
    text,tables,warns=extract_text_from_docx(fb)
    fmap,q,defs=get_faculty_map(text,tables)
    slots=extract_busy_slots_from_docx(tables)
    if not slots: slots=extract_busy_slots_from_text(text)
    det=sorted(set(fmap.keys())|{s["faculty_code"] for s in slots})
    multi=detect_multiple_batches(text)
    return {"faculty_map":fmap,"busy_slots":slots,"detected_codes":det,"total_slots":len(slots),
            "warnings":warns,"used_defaults":defs,"parse_quality":"extracted" if q=="extracted_docx" else "partial",
            "map_quality":q,"multi_batch":multi}

def _tt_xlsx(fb):
    text,rows,warns=extract_text_from_xlsx(fb)
    fmap,q,defs=get_faculty_map(text,[rows] if rows else [])
    slots=extract_busy_slots_from_text(text)
    multi=detect_multiple_batches(text)
    return {"faculty_map":fmap,"busy_slots":slots,"detected_codes":sorted(fmap.keys()),
            "total_slots":len(slots),"warnings":warns,"used_defaults":defs,
            "parse_quality":"partial","map_quality":q,"multi_batch":multi}

def _tt_pdf(fb):
    text,warns=extract_text_from_pdf(fb)
    fmap,q,defs=get_faculty_map(text)
    slots=extract_busy_slots_from_text(text)
    det=sorted(set(fmap.keys())|{s["faculty_code"] for s in slots})
    multi=detect_multiple_batches(text)
    return {"faculty_map":fmap,"busy_slots":slots,"detected_codes":det,"total_slots":len(slots),
            "warnings":warns,"used_defaults":defs,"parse_quality":q,"map_quality":q,"multi_batch":multi}

def _tt_csv(fb):
    try:
        text=fb.decode("utf-8-sig"); reader=csvlib.DictReader(io.StringIO(text)); rows=list(reader)
    except Exception as e:
        return {"faculty_map":dict(KNOWN_FACULTY),"busy_slots":[],"detected_codes":[],
                "total_slots":0,"warnings":[str(e)],"used_defaults":["CSV error — built-in BMS list"],
                "parse_quality":"fallback","multi_batch":False}
    fmap,slots={},[]; multi=False
    for row in rows:
        n={k.strip().lower():(v or "").strip() for k,v in row.items() if k}
        code=(n.get("faculty_code") or n.get("code") or "").upper()
        name=n.get("faculty_name") or n.get("name") or ""
        if code and name: fmap[code]=name
        if code and n.get("day",""):
            slots.append({"faculty_code":code,"subject":n.get("subject",""),
                          "day":n.get("day",""),"time_slot":"","room":""})
        if "batch" in n.get("batch","").lower() or "sem" in n.get("semester","").lower(): multi=True
    if not fmap: fmap=dict(KNOWN_FACULTY)
    return {"faculty_map":fmap,"busy_slots":slots,"detected_codes":sorted(fmap.keys()),
            "total_slots":len(slots),"warnings":[],"used_defaults":[],"parse_quality":"good",
            "map_quality":"csv","multi_batch":multi}

def parse_exam_schedule(file_bytes, filename=None, custom_rooms=None)->Dict:
    # Use custom rooms if provided
    if custom_rooms is not None:
        rooms = custom_rooms
    else:
        rooms = DEFAULT_ROOMS()  # or whatever your fallback is
    if not file_bytes:
        pl=_make_placeholders()
        return {"exams":pl,"variant":"unknown","rooms":deepcopy(DEFAULT_ROOMS),
                "warnings":["Empty file"],"used_defaults":["Placeholders created"],
                "total":len(pl),"parse_quality":"fallback"}
    ext=filename.lower().rsplit(".",1)[-1] if "." in filename else "pdf"
    if ext=="csv": return _ex_csv(file_bytes)
    if ext in("docx","doc"): text,_,warns=extract_text_from_docx(file_bytes)
    elif ext in("xlsx","xls","ods"): text,_,warns=extract_text_from_xlsx(file_bytes)
    else: text,warns=extract_text_from_pdf(file_bytes)
    return _ex_text(text,warns)

def _ex_text(text,extra_warns=None):
    warns=list(extra_warns or []); used=[]
    tl=text.lower()
    variant=("end_sem" if any(k in tl for k in["end main examination","end semester","mmc10","mmc20"])
             else "lab_cie" if any(k in tl for k in["lab cie","unix lab","batches"]) else "theory_cie")
    rooms=extract_rooms(text); total=sum(r["students"] for r in rooms)
    if rooms==deepcopy(DEFAULT_ROOMS): used.append("Rooms not found — using default CA1/CA2/BT1/BT2")
    gs,ge=find_time_range(text)
    cs=gs or("2:00 PM" if variant=="end_sem" else "9:15 AM")
    ce=ge or("5:00 PM" if variant=="end_sem" else "10:30 AM")
    if not gs: used.append(f"Exam time not found — using default {cs}–{ce}")
    exams=[]; seen=set(); cur_dt=None; cur_date=""; lines=text.splitlines()
    for i,raw in enumerate(lines):
        line=re.sub(r'\s{2,}',' ',raw.strip())
        if not line: continue
        for pat in[r'\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b',
                   r'\b(\d{1,2}[\s\-](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[\s\-]\d{4})\b']:
            dm=re.search(pat,line,re.IGNORECASE)
            if dm:
                dt=parse_date(dm.group(1))
                if dt:
                    cur_dt,cur_date=dt,fmt_date(dt)
                    s,e=find_time_range(line)
                    if s: cs,ce=s,e
                break
        s,e=find_time_range(line)
        if s: cs,ce=s,e
        window=line+(" "+lines[i+1].strip() if i+1<len(lines) else "")
        subj=detect_subject(window)
        if subj and cur_date:
            code=""
            for nb in lines[max(0,i-2):i+3]:
                mc=re.match(r'^([A-Z]{2,4}\d{3})\s*$',nb.strip())
                if mc: code=mc.group(1); break
            key=(subj,cur_date)
            if key not in seen:
                seen.add(key)
                day=cur_dt.strftime("%A") if cur_dt else ""
                etype={"end_sem":"End Semester","lab_cie":"Lab CIE","theory_cie":"Theory CIE"}.get(variant,"CIE")
                exams.append({"exam_type":etype,"course_code":code,"subject":subj,"section":"All sections",
                               "exam_date":cur_date,"day":day,"start_time":cs,"end_time":ce,
                               "total_students":total,"rooms":rooms,"used_defaults":[]})
    if not exams:
        exams=_make_placeholders(rooms=rooms,total_students=total)
        used.append("No exam slots extracted — created placeholders"); warns.append("Try uploading as CSV")
    return {"exams":exams,"variant":variant,"rooms":rooms,"warnings":warns,"used_defaults":used,
            "total":len(exams),"parse_quality":"good" if not used else "partial"}

def _ex_csv(fb):
    try:
        text=fb.decode("utf-8-sig"); reader=csvlib.DictReader(io.StringIO(text)); rows=list(reader)
    except Exception as e:
        pl=_make_placeholders()
        return {"exams":pl,"variant":"csv","rooms":deepcopy(DEFAULT_ROOMS),"warnings":[str(e)],
                "used_defaults":["CSV error — placeholders"],"total":len(pl),"parse_quality":"fallback"}
    exams=[]; warns=[]
    for i,row in enumerate(rows,1):
        n={k.strip().lower():(v or "").strip() for k,v in row.items() if k}
        subj=n.get("subject") or n.get("course_name") or ""
        if not subj: warns.append(f"Row {i}: no subject"); continue
        dt=parse_date(n.get("exam_date") or n.get("date") or "")
        if not dt: warns.append(f"Row {i}: bad date"); continue
        st=fmt_time(parse_time(n.get("start_time","") or n.get("start",""))) or "9:15 AM"
        et=fmt_time(parse_time(n.get("end_time","") or n.get("end",""))) or "10:30 AM"
        try: total=int(n.get("total_students") or n.get("students") or DEFAULT_STUDENTS)
        except: total=DEFAULT_STUDENTS
        rooms=[]
        for j in range(1,6):
            rn=n.get(f"room{j}",""); sc=n.get(f"students{j}","")
            if rn: rooms.append({"room":rn,"students":int(sc) if sc.isdigit() else 30,"usn_range":"","usn_start":"","usn_end":""})
        if not rooms: rooms=deepcopy(DEFAULT_ROOMS)
        exams.append({"exam_type":"CIE","course_code":n.get("course_code",""),"subject":subj,
                      "section":n.get("section","All sections"),"exam_date":fmt_date(dt),
                      "day":dt.strftime("%A"),"start_time":st,"end_time":et,
                      "total_students":total,"rooms":rooms,"used_defaults":[]})
    if not exams: exams=_make_placeholders(); warns.append("No valid rows — placeholders created")
    return {"exams":exams,"variant":"csv","rooms":deepcopy(DEFAULT_ROOMS),"warnings":warns,
            "used_defaults":[],"total":len(exams),"parse_quality":"good"}

def _make_placeholders(rooms=None,total_students=DEFAULT_STUDENTS):
    if rooms is None: rooms=deepcopy(DEFAULT_ROOMS)
    today=datetime.now(); entries=[]
    for off,subj in enumerate(KNOWN_SUBJECTS[:4]):
        dt=today+timedelta(days=off+1)
        entries.append({"exam_type":"Theory CIE (placeholder)","course_code":"","subject":subj,
                        "section":"All sections","exam_date":fmt_date(dt),"day":dt.strftime("%A"),
                        "start_time":"9:15 AM","end_time":"10:30 AM","total_students":total_students,
                        "rooms":rooms,"is_placeholder":True,"used_defaults":["Placeholder"]})
    return entries